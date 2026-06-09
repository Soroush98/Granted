"""CIHR Grants and Awards ingestion (the one source with scientific abstracts).

CIHR publishes its full funding history — including a free-text scientific
abstract per project — on the Open Government portal:

    https://open.canada.ca/data/en/dataset/49edb1d7-5cb4-4fa7-897c-515d1aad5da3

One amalgamated XLSX per fiscal year (since 2000-01), updated quarterly. Each
file has five tabs; we read **G&A_S&B** ("Grants & Awards"), one row per funded
project, with these columns (bilingual EN_FR header names):

    FundingReferenceNumber_… ........ stable per-project id (our award_id)
    FamilyName_… / FirstName_… ...... the nominated principal investigator
    ResearchInstitutionNameEN_… ..... host institution (PI's lab)
    InstitutionPaidNameEN_… ......... fallback institution (who got the cheque)
    ProgramNameEN_… ................. e.g. "Project Grant", "Operating Grant"
    ProgramTypeEN_… ................. "Grant Program" (PI-held) vs "Award
                                       Program" (trainee: doctoral/Banting/etc.)
    TotalAmountAwarded_… ............ full award value
    FundingStartDate_… / …EndDate_… . ISO dates
    ApplicationTitle_… .............. project title
    ApplicationAbstract_… ........... the scientific abstract (~96% populated)
    ApplicationKeywords_… ........... keywords

Why this source matters: the federal Proactive-Disclosure feed (proactive.py)
also carries CIHR rows, but only the admin/program text — so a query like
"neuroradiology brain MRI" matched almost no CIHR there. These abstracts are the
real scientific descriptions, so health-research PIs become retrievable.

We tag everything under the `CIHR` program code (auto-created if missing) and
store the host institution as a `university` recipient — the same academic
cluster NSERC/FRQ feed, so the "find a PI" flow surfaces them. The raw row is
persisted verbatim so the web app's holder/funder extraction can read the
PI name, program type, and "CIHR" source straight from these columns.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import openpyxl
from rich.console import Console

from .. import db
from ..normalize import normalize_org_name

console = Console()

_GA_SHEET = "G&A_S&B"
_DEFAULT_ORG_TYPE = "university"

# Column header names on the G&A tab (bilingual EN_FR).
_C_REF = "FundingReferenceNumber_NumeroReferenceFinancement"
_C_FAMILY = "FamilyName_NomFamille"
_C_FIRST = "FirstName_Prenom"
_C_RESEARCH_INST = "ResearchInstitutionNameEN_NomEtablissementRechercheAN"
_C_PAID_INST = "InstitutionPaidNameEN_NomEtablissementPayeAN"
_C_PROGRAM = "ProgramNameEN_NomProgrammeAN"
_C_PROGRAM_TYPE = "ProgramTypeEN_TypeProgrammeAN"
_C_AMOUNT = "TotalAmountAwarded_MontantTotalAccorde"
_C_START = "FundingStartDate_DatePremierVersement"
_C_END = "FundingEndDate_DateDernierVersement"
_C_FY = "FiscalYear_AnneeFinanciere"
_C_TITLE = "ApplicationTitle_TitreDemande"
_C_ABSTRACT = "ApplicationAbstract_ResumeDemande"
_C_KEYWORDS = "ApplicationKeywords_MotsClesDemande"


def _cell(row: dict[str, Any], key: str) -> str:
    v = row.get(key)
    return "" if v is None else str(v).strip()


def _amount(row: dict[str, Any]) -> float | None:
    raw = _cell(row, _C_AMOUNT).replace(",", "").replace("$", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _date(row: dict[str, Any], key: str) -> str | None:
    v = row.get(key)
    if v is None:
        return None
    # openpyxl hands back either an ISO-ish string or a datetime; normalize to
    # the YYYY-MM-DD date string the grants table expects.
    s = str(v).strip()
    return s[:10] or None if s else None


def _org_name(row: dict[str, Any]) -> str:
    # The PI's research institution is what we want (their lab's home); fall back
    # to the paying institution when CIHR leaves the research one blank.
    return _cell(row, _C_RESEARCH_INST) or _cell(row, _C_PAID_INST)


def _description(row: dict[str, Any]) -> str | None:
    """The scientific abstract, with keywords appended so they're searchable."""
    abstract = _cell(row, _C_ABSTRACT)
    keywords = _cell(row, _C_KEYWORDS)
    if abstract and keywords:
        return f"{abstract}\n\nKeywords: {keywords}"
    return abstract or keywords or None


def ingest_xlsx(path: Path, *, limit: int | None = None) -> tuple[int, int]:
    """Stream one CIHR fiscal-year XLSX into the DB.

    Reads the G&A tab row-by-row (read-only mode keeps the 12 MB+ files off the
    heap), skipping rows with no host institution. Re-ingesting later fiscal-year
    files updates the same grant — the (program_id, FundingReferenceNumber) key
    dedupes the recurrences of a multi-year project. Returns (companies_seen,
    grants_inserted)."""
    program_id = db.ensure_program_id(
        "CIHR",
        name="CIHR Grants and Awards",
        agency="CIHR",
    )
    seen_companies: set[str] = set()
    grants_inserted = 0

    wb = openpyxl.load_workbook(path, read_only=True)
    if _GA_SHEET not in wb.sheetnames:
        raise RuntimeError(
            f"{path.name}: no {_GA_SHEET!r} sheet (found {wb.sheetnames}). "
            "Is this a CIHR Grants & Awards file?"
        )
    ws = wb[_GA_SHEET]
    row_iter = ws.iter_rows(values_only=True)
    header = list(next(row_iter, []) or [])

    def flush(rows: list[dict[str, Any]]) -> int:
        prepared: list[dict[str, Any]] = []
        company_payload: list[dict[str, Any]] = []
        for row in rows:
            org_name = _org_name(row)
            if not org_name:
                continue
            normalized = normalize_org_name(org_name)
            if not normalized:
                continue
            ref = _cell(row, _C_REF)
            if not ref:
                continue
            company_payload.append(
                {
                    "display_name": org_name,
                    "normalized_name": normalized,
                    "org_type": _DEFAULT_ORG_TYPE,
                }
            )
            prepared.append(
                {
                    "row": row,
                    "normalized": normalized,
                    "award_id": ref,
                    "title": _cell(row, _C_TITLE) or None,
                    "description": _description(row),
                }
            )
        if not prepared:
            return 0

        company_map = db.bulk_upsert_companies(company_payload)
        seen_companies.update(company_map.values())

        grant_payload: list[dict[str, Any]] = []
        for p in prepared:
            cid = company_map.get((p["normalized"], _DEFAULT_ORG_TYPE))
            if not cid:
                continue
            p["company_id"] = cid
            grant_payload.append(
                {
                    "program_id": program_id,
                    "recipient_id": cid,
                    "title": p["title"],
                    "description": p["description"],
                    "amount_cad": _amount(p["row"]),
                    "start_date": _date(p["row"], _C_START),
                    "end_date": _date(p["row"], _C_END),
                    "fiscal_year": _cell(p["row"], _C_FY) or None,
                    "award_id": p["award_id"],
                    "raw": p["row"],
                }
            )
        grant_map = db.bulk_upsert_grants(grant_payload)

        chunk_payload: list[dict[str, Any]] = []
        for p in prepared:
            if "company_id" not in p:
                continue
            gid = grant_map.get((program_id, p["award_id"]))
            if not gid:
                continue
            if p["title"]:
                chunk_payload.append(
                    {
                        "grant_id": gid,
                        "company_id": p["company_id"],
                        "kind": "grant_title",
                        "content": p["title"],
                    }
                )
            if p["description"]:
                chunk_payload.append(
                    {
                        "grant_id": gid,
                        "company_id": p["company_id"],
                        "kind": "grant_description",
                        "content": p["description"][:2000],
                    }
                )
        # Defer embedding — embed-pending vectorizes new chunks later in small
        # batches so HNSW maintenance doesn't trip Supabase's statement timeout.
        db.bulk_insert_chunks(chunk_payload, embed=False)
        return len(grant_payload)

    batch: list[dict[str, Any]] = []
    scanned = 0
    for values in row_iter:
        if limit is not None and scanned >= limit:
            break
        scanned += 1
        batch.append({h: v for h, v in zip(header, values)})
        if len(batch) >= 400:
            grants_inserted += flush(batch)
            batch = []
            console.log(f"CIHR: scanned {scanned}, {grants_inserted} grants so far…")
    if batch:
        grants_inserted += flush(batch)

    wb.close()
    return len(seen_companies), grants_inserted
