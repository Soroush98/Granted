"""NHMRC (National Health and Medical Research Council) ingestion.

NHMRC is Australia's medical-research funder — the ARC's health counterpart and
the right source for a clinician/MD looking for an Australian lab. It publishes
"Research Funding Datasets" (one XLSX covering all grants since 1990, refreshed
annually) including a plain-language description per grant, the chief
investigator, and the administering institution.

    https://www.nhmrc.gov.au/funding/outcomes-and-data-research/research-funding-statistics-and-data

Download the dataset XLSX from that page (it sits behind a tokenised link, so it
can't be fetched headlessly) and pass the local path to `ingest_xlsx`.

NHMRC has reorganised these files several times, so column headers vary between
vintages ("CIA Name" vs "Chief Investigator A", "Plain Description" vs "Media
Summary", etc.). Rather than hard-code one spelling, we DETECT each column by
fuzzy-matching the header row and log the mapping on first run. We persist a few
normalised markers (`_funder`, `_pi`, `_program`) into the stored raw so the web
app's holder extraction doesn't depend on the exact source spelling.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any

import openpyxl
from rich.console import Console

from .. import db
from ..normalize import normalize_org_name

console = Console()

_DEFAULT_ORG_TYPE = "university"

# For each logical field, ordered regexes tried against normalised headers
# (lowercased, non-alphanumerics collapsed to spaces). First match wins; earlier
# patterns are more specific so they beat looser fallbacks.
_FIELD_PATTERNS: dict[str, list[str]] = {
    "pi": [r"\bcia name\b", r"chief investigator a\b", r"\bcia\b", r"chief investigator", r"investigator name", r"lead investigator"],
    "institution": [r"admin(istering)? (institution|organisation)", r"\binstitution\b", r"\borganisation\b"],
    "title": [r"grant title", r"application title", r"project title", r"\btitle\b"],
    "abstract": [r"plain( language)? description", r"plain( language)? summary", r"media summary", r"lay summary", r"\bsummary\b", r"\babstract\b", r"description"],
    "program": [r"grant sub.?type", r"grant type", r"\bscheme\b", r"sub.?type", r"\btype\b"],
    "amount": [r"total (amount|funding|paid)", r"amount awarded", r"\bamount\b", r"\bfunding\b", r"\btotal\b"],
    "start": [r"start year", r"commenc", r"start date", r"funding commenc"],
    "end": [r"end year", r"end date", r"completion"],
    "state": [r"\bstate\b", r"province"],
    "id": [r"application id", r"grant id", r"app id", r"\bid\b", r"reference"],
}


def _norm(h: Any) -> str:
    return re.sub(r"[^a-z0-9]+", " ", str(h or "").lower()).strip()


def _detect_columns(header: list[Any]) -> dict[str, int]:
    """Map each logical field to a column index using the pattern table."""
    normed = [_norm(h) for h in header]
    mapping: dict[str, int] = {}
    for field, patterns in _FIELD_PATTERNS.items():
        for pat in patterns:
            hit = next((i for i, h in enumerate(normed) if re.search(pat, h)), None)
            if hit is not None:
                mapping[field] = hit
                break
    return mapping


def _pick_sheet(wb: openpyxl.Workbook) -> tuple[str, list[Any], dict[str, int]]:
    """Pick the sheet whose header row maps the most of our fields (a data tab,
    not a glossary). Returns (sheet_name, header, column_mapping)."""
    best: tuple[int, str, list[Any], dict[str, int]] | None = None
    for name in wb.sheetnames:
        ws = wb[name]
        header = next(ws.iter_rows(values_only=True, max_row=1), None)
        if not header:
            continue
        mapping = _detect_columns(list(header))
        score = len(mapping)
        if best is None or score > best[0]:
            best = (score, name, list(header), mapping)
    if best is None:
        raise RuntimeError("no readable sheet found in NHMRC workbook")
    _, name, header, mapping = best
    return name, header, mapping


def _val(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> str:
    i = mapping.get(field)
    if i is None or i >= len(row):
        return ""
    v = row[i]
    return "" if v is None else str(v).strip()


def _amount(row: tuple[Any, ...], mapping: dict[str, int]) -> float | None:
    raw = _val(row, mapping, "amount").replace(",", "").replace("$", "")
    if not raw:
        return None
    try:
        return float(raw)
    except ValueError:
        return None


def _year_date(value: str, *, end: bool) -> str | None:
    """Coerce a year or date cell to an ISO date. A bare year → Jan 1 (start) or
    Dec 31 (end)."""
    v = value.strip()
    if not v:
        return None
    if re.fullmatch(r"\d{4}", v):
        return f"{v}-12-31" if end else f"{v}-01-01"
    return v[:10]


def ingest_xlsx(path: Path, *, limit: int | None = None) -> tuple[int, int]:
    """Stream an NHMRC research-funding XLSX into the DB. Returns
    (institutions_seen, grants_inserted)."""
    program_id = db.ensure_program_id(
        "NHMRC",
        name="NHMRC Research Funding",
        agency="NHMRC",
    )
    wb = openpyxl.load_workbook(path, read_only=True)
    sheet_name, header, mapping = _pick_sheet(wb)
    console.log(f"NHMRC: using sheet {sheet_name!r}; detected columns:")
    for field, i in sorted(mapping.items()):
        console.log(f"    {field:12} -> col[{i}] {header[i]!r}")
    missing = [f for f in ("pi", "institution", "abstract", "id") if f not in mapping]
    if missing:
        console.log(f"[yellow]NHMRC: WARNING missing columns {missing} — check the header names.[/]")

    ws = wb[sheet_name]
    row_iter = ws.iter_rows(values_only=True)
    next(row_iter, None)  # skip header

    seen_companies: set[str] = set()
    grants_inserted = 0

    def flush(rows: list[tuple[Any, ...]]) -> int:
        prepared: list[dict[str, Any]] = []
        company_payload: list[dict[str, Any]] = []
        for row in rows:
            org_name = _val(row, mapping, "institution")
            if not org_name:
                continue
            normalized = normalize_org_name(org_name)
            if not normalized:
                continue
            award_id = _val(row, mapping, "id")
            if not award_id:
                continue
            pi = _val(row, mapping, "pi")
            program = _val(row, mapping, "program")
            abstract = _val(row, mapping, "abstract")
            title = _val(row, mapping, "title")
            # Persist original cells plus stable markers the web app keys off.
            raw = {str(h): (row[i] if i < len(row) else None) for i, h in enumerate(header)}
            raw["_funder"] = "NHMRC"
            raw["_pi"] = pi
            raw["_program"] = program
            company_payload.append(
                {
                    "display_name": org_name,
                    "normalized_name": normalized,
                    "org_type": _DEFAULT_ORG_TYPE,
                    "province": _val(row, mapping, "state") or None,
                }
            )
            prepared.append(
                {
                    "row": row,
                    "raw": raw,
                    "normalized": normalized,
                    "award_id": award_id,
                    "title": title or None,
                    "description": abstract or None,
                }
            )
        if not prepared:
            return 0

        company_map = db.bulk_upsert_companies(company_payload)
        seen_companies.update(company_map.values())

        grant_payload: list[dict[str, Any]] = []
        for p in prepared:
            cid = company_map.get((p["normalized"], _DEFAULT_ORG_TYPE))
            if not cid:
                continue
            p["company_id"] = cid
            grant_payload.append(
                {
                    "program_id": program_id,
                    "recipient_id": cid,
                    "title": p["title"],
                    "description": p["description"],
                    "amount_cad": _amount(p["row"], mapping),
                    "start_date": _year_date(_val(p["row"], mapping, "start"), end=False),
                    "end_date": _year_date(_val(p["row"], mapping, "end"), end=True),
                    "award_id": p["award_id"],
                    "raw": p["raw"],
                }
            )
        grant_map = db.bulk_upsert_grants(grant_payload)

        chunk_payload: list[dict[str, Any]] = []
        for p in prepared:
            if "company_id" not in p:
                continue
            gid = grant_map.get((program_id, p["award_id"]))
            if not gid:
                continue
            if p["title"]:
                chunk_payload.append(
                    {"grant_id": gid, "company_id": p["company_id"], "kind": "grant_title", "content": p["title"]}
                )
            if p["description"]:
                chunk_payload.append(
                    {"grant_id": gid, "company_id": p["company_id"], "kind": "grant_description", "content": p["description"][:2000]}
                )
        db.bulk_insert_chunks(chunk_payload, embed=False)
        return len(grant_payload)

    batch: list[tuple[Any, ...]] = []
    scanned = 0
    for row in row_iter:
        if limit is not None and scanned >= limit:
            break
        scanned += 1
        batch.append(row)
        if len(batch) >= 400:
            grants_inserted += flush(batch)
            batch = []
            console.log(f"NHMRC: scanned {scanned}, {grants_inserted} grants so far…")
    if batch:
        grants_inserted += flush(batch)

    wb.close()
    return len(seen_companies), grants_inserted
